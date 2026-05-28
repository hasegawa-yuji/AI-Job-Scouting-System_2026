import os
import json
import re
import time
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==========================================
# 1. 初期設定
# ==========================================

# 【利用方法】環境変数にキーを設定するか、下のクォーテーション内に直接APIキーを記述してください
API_KEY = os.getenv("OPENAI_API_KEY") or "YOUR_OPENAI_API_KEY_HERE"

EXCEL_FILE = "求人精査/2026.xlsx"
OUTPUT_FILE = "求人精査/査定結果_完全版.xlsx"

# OpenAIクライアント初期化（timeout=60秒）
client = OpenAI(
    api_key=API_KEY,
    timeout=60
)

# 最大解析数
TOTAL_MAX_TEST = 9999

# ==========================================
# 2. Excel読み込み
# ==========================================

print("Excelファイルを読み込み中...")

all_sheets = pd.read_excel(
    EXCEL_FILE,
    sheet_name=None,
    header=None
)

# ==========================================
# 3. AI評価プロンプト（汎用テンプレート版）
# ==========================================

PROMPT_TEMPLATE = """
あなたは転職市場分析AIです。

応募者スペックと求人情報を比較し、
客観的かつ実務的に評価してください。

感情的配慮や一般論は禁止です。
転職市場に基づき現実的に評価してください。

# 最優先方針
1. 通過可能性
2. 長期的に消耗しにくいか
3. 業務負荷
4. 初年度年収の現実性
5. 将来性

# 応募者スペック（※利用時にご自身のスペックに書き換えてください）
- 年齢 / 学歴: 【例: 30代後半 / 高卒】
- 主要経験1: 【例: 社内情シス、IT資産管理、ヘルプデスク、キッティング】
- 主要経験2: 【例: RPA設計構築（単独での業務整理・自動化実装）】
- 主要経験3: 【例: 現場経験（製造・物流などでの現場改善・作業標準化）】
- スキル補足: 【例: VBA/PythonはAI補助前提であり、プログラミング専門職レベルではない】

# 高評価になりやすい求人（※希望の職種や業界に書き換えてください）
- 【例: 社内SE、情シス、IT運用管理、製造業IT担当、業務改善、DX補助】
- 相性の良い企業: 【例: 現場改善文化がある会社、昔ながらの現場文化が残る中小〜中堅企業】

# 減点・警戒条件
- 【例: 高度開発必須、PL/PM必須、AWS/GCP上流設計専任、コンサル寄り】
- 【例: 激務リスク（24h対応、夜間障害、1人情シス、多重下請けSES）】
- 【例: 他県への数ヶ月単位の長期研修リスク】

# 危険ワード（見つけたら減点対象とするキーワード）
- 【例: 少数精鋭、裁量、スピード感、成長環境、DX推進、主体性、自走、0→1、マルチタスク】

# 求人情報
{}

# 出力ルール
JSONのみ出力してください。
説明は禁止。
markdown禁止。

# 出力形式
{{
  "応募ランク": "",
  "応募判断": "",
  "通過率": "",
  "年収現実性": "",
  "業務負荷": "",
  "精神的負荷": "",
  "属人化リスク": "",
  "問い合わせ地獄リスク": "",
  "総評": "",
  "懸念点": [],
  "総合評価": "",
  "総合点": 0
}}
"""

# ==========================================
# 4. 初期化
# ==========================================

structured_results = []

print("\n--- AI求人精査を開始します ---")

job_count = 0

company_keywords = [
    "株式会社",
    "有限会社",
    "合同会社"
]

# 抽出対象キーワード（ノイズ削減用）
filter_keywords = [
    "年収",
    "給与",
    "休日",
    "勤務地",
    "業務内容",
    "仕事内容",
    "応募条件",
    "必須",
    "歓迎",
    "社内SE",
    "情シス",
    "IT",
    "DX",
    "運用",
    "ヘルプデスク"
]

processed_companies = set()

# ==========================================
# 5. シート巡回
# ==========================================

for sheet_name, df in all_sheets.items():

    if job_count >= TOTAL_MAX_TEST:
        break

    print(f"\nシート解析中: {sheet_name}")

    for col_idx in range(df.shape[1]):

        if job_count >= TOTAL_MAX_TEST:
            break

        company_name = None

        # 企業名探索
        for row_idx in range(df.shape[0]):

            cell_val = df.iloc[row_idx, col_idx]

            if pd.notna(cell_val):

                cell_str = str(cell_val).strip()

                if any(
                    keyword in cell_str
                    for keyword in company_keywords
                ):
                    company_name = cell_str
                    break

        if not company_name:
            continue

        # 重複除外
        if company_name in processed_companies:
            continue

        processed_companies.add(company_name)

        print(f"求人発見: {company_name}")

        # ==========================================
        # 求人本文抽出（キーワード抽出 ＆ フォールバック機能）
        # ==========================================

        full_lines = []
        filtered_lines = []

        for row_idx in range(df.shape[0]):

            val = df.iloc[row_idx, col_idx]

            if pd.notna(val):

                text = str(val).strip()

                if text != "":

                    clean_text = " ".join(text.split())

                    full_lines.append(clean_text)

                    if any(
                        kw in clean_text
                        for kw in filter_keywords
                    ):
                        filtered_lines.append(clean_text)

        filtered_text = "\n".join(filtered_lines)

        # フォールバック判定（文字数または行数が少なすぎる場合は全文モードへ）
        if len(filtered_text) < 500 or len(filtered_lines) < 5:
            job_text = "\n".join(full_lines)
        else:
            job_text = filtered_text

        # 長文対策（上限8000文字カット）
        if len(job_text) > 8000:
            job_text = job_text[:8000]

        # ==========================================
        # OpenAI API 呼び出し
        # ==========================================

        try:

            response = client.chat.completions.create(
                model="gpt-4o-mini",
                response_format={"type": "json_object"},
                messages=[
                    {
                        "role": "user",
                        "content": PROMPT_TEMPLATE.format(job_text)
                    }
                ],
                temperature=0.1
            )

            ai_output = response.choices[0].message.content.strip()

            # markdown除去
            if ai_output.startswith("```"):
                ai_output = re.sub(r"^```[a-zA-Z]*\n", "", ai_output)
                ai_output = re.sub(r"\n```$", "", ai_output)
            ai_output = ai_output.strip()

            # ==========================================
            # JSON解析
            # ==========================================

            try:
                job_data = json.loads(ai_output)
            except Exception:
                print(f"\nJSON解析失敗: {company_name}")
                print("\n--- AI生出力ここから ---")
                print(ai_output)
                print("--- AI生出力ここまで ---\n")

                structured_results.append({
                    "会社名": company_name,
                    "応募ランク": "JSON_ERROR",
                    "応募判断": "JSON解析失敗",
                    "総合点": 0
                })
                continue

            # ==========================================
            # Excel用データ整形
            # ==========================================

            raw_kn = job_data.get("懸念点", [])
            if isinstance(raw_kn, list):
                string_kn = ", ".join(raw_kn)
            else:
                string_kn = str(raw_kn)

            formatted_data = {
                "会社名": company_name,
                "応募ランク": job_data.get("応募ランク", ""),
                "応募判断": job_data.get("応募判断", ""),
                "総合点": job_data.get("総合点", 0),
                "通過率": job_data.get("通過率", ""),
                "年収現実性": job_data.get("年収現実性", ""),
                "業務負荷": job_data.get("業務負荷", ""),
                "精神的負荷": job_data.get("精神的負荷", ""),
                "属人化リスク": job_data.get("属人化リスク", ""),
                "問い合わせ地獄リスク": job_data.get("問い合わせ地獄リスク", ""),
                "総評": job_data.get("総評", ""),
                "懸念点": string_kn,
                "総合評価": job_data.get("総合評価", "")
            }

            structured_results.append(formatted_data)
            job_count += 1
            print(f"解析完了 ({job_count}件)")

        except Exception as e:
            print(f"\nAPIエラー: {company_name}")
            print(str(e))
            structured_results.append({
                "会社名": company_name,
                "応募ランク": "API_ERROR",
                "応募判断": str(e),
                "総合点": 0
            })

        finally:
            # 429 Rate Limit 対策（1秒待機）
            time.sleep(1)

# ==========================================
# 6. Excel出力
# ==========================================

if structured_results:

    output_df = pd.DataFrame(structured_results)

    # 総合点ソート
    if "総合点" in output_df.columns:
        output_df["総合点"] = pd.to_numeric(output_df["総合点"], errors="coerce").fillna(0).astype(int)
        output_df = output_df.sort_values(by="総合点", ascending=False)

    # 列順固定
    columns_order = [
        "会社名", "応募ランク", "応募判断", "総合点", "通過率", "年収現実性",
        "業務負荷", "精神的負荷", "属人化リスク", "問い合わせ地獄リスク", "総評", "懸念点", "総合評価"
    ]
    output_df = output_df.reindex(columns=columns_order)

    # Excel保存
    output_df.to_excel(OUTPUT_FILE, index=False)
    print("\nExcel保存完了")

    # ==========================================
    # Excel装飾（openpyxlによる色付け）
    # ==========================================

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        headers[cell.value] = idx

    for row in ws.iter_rows(min_row=2):
        apply_value = row[headers["応募判断"] - 1].value
        score_value = row[headers["総合点"] - 1].value

        # スコアによる色付け
        if isinstance(score_value, (int, float)):
            if score_value >= 85:
                for cell in row:
                    cell.fill = green_fill
            elif score_value <= 40:
                for cell in row:
                    cell.fill = red_fill

        # 「即応募」の最優先色付け
        if apply_value == "即応募":
            for cell in row:
                cell.fill = yellow_fill

    wb.save(OUTPUT_FILE)

    print("\n===================================")
    print("すべての処理が完了しました")
    print(f"解析件数: {job_count}")
    print(f"保存先: {OUTPUT_FILE}")
    print("===================================")

else:
    print("求人データが見つかりませんでした")